import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_report_directory():
    # Create the reports directory if it doesn't exist
    if not os.path.exists("reports"):
        os.makedirs("reports")

def read_existing_reports(filename="reports/test_report.xlsx"):
    """Read the existing Excel file and return a list of rows (test results)."""
    existing_rows = []
    if os.path.isfile(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        
        # Skip the header row and read data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            existing_rows.append(list(row))
        
        workbook.close()
    return existing_rows

def write_to_xlsx(test_data, filename="reports/test_report.xlsx"):
    create_report_directory()  # Ensure the reports directory exists

    # Read existing reports to check if the test result already exists
    existing_reports = read_existing_reports(filename)

    try:
        # Check if file exists, if not create a new workbook
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Set up headers with styling
            headers = ['Page URL', 'Test Case', 'Status', 'Comments']
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

        # Flag to track if we updated an existing test result
        updated = False
        
        # Find the last row with data
        last_row = sheet.max_row

        # Iterate through existing rows to update or preserve
        for row_idx in range(2, last_row + 1):
            existing_url = sheet.cell(row=row_idx, column=1).value
            existing_test_case = sheet.cell(row=row_idx, column=2).value
            
            # Check if this row matches the current test
            if existing_url == test_data[0] and existing_test_case == test_data[1]:
                # Update the existing row
                for col, value in enumerate(test_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col, value=value)
                    
                    # Color-code status
                    if col == 3:
                        if value == 'PASS':
                            cell.font = Font(color="006400")  # Dark Green
                        elif value == 'FAIL':
                            cell.font = Font(color="8B0000")  # Dark Red
                
                updated = True
                break

        # If test result didn't exist, append new row
        if not updated:
            last_row += 1
            for col, value in enumerate(test_data, start=1):
                cell = sheet.cell(row=last_row, column=col, value=value)
                
                # Color-code status
                if col == 3:
                    if value == 'PASS':
                        cell.font = Font(color="006400")  # Dark Green
                    elif value == 'FAIL':
                        cell.font = Font(color="8B0000")  # Dark Red

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            sheet.column_dimensions[column].width = max_length + 2

        # Save the workbook
        workbook.save(filename)
        workbook.close()

    except Exception as e:
        print(f"Error writing to Excel: {e}")

def run_test(driver, url, test_func, test_name):
    try:
        driver.get(url)
        result, comments = test_func(driver)
        status = 'PASS' if result else 'FAIL'
        write_to_xlsx([url, test_name, status, comments])
    except Exception as e:
        write_to_xlsx([url, test_name, 'FAIL', str(e)])